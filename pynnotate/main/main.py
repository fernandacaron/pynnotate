import queue

error_queue = queue.Queue()

def load_synonyms_from_str(input_str, root=None):
	import yaml

	try:
		syn_dict = yaml.safe_load(input_str)
		if not isinstance(syn_dict, dict):
			report_message("Parsed synonyms must be a dictionary.", "error", "Error", root)
			return {}
		return syn_dict
	except Exception as e:
		report_message(f"Failed to parse synonyms: {e}", "error", "Error", root)
		return {}
		
def merge_synonyms(base_dict, user_dict):
    merged = base_dict.copy()
    for key, synonyms in user_dict.items():
        if isinstance(synonyms, str):
            synonyms = [synonyms]
        elif not isinstance(synonyms, list):
            synonyms = list(synonyms) if synonyms else []
        
        if key in merged:
            merged[key] = list(set(merged[key] + synonyms))
        else:
            merged[key] = synonyms
    return merged


def normalize(text):
	import unicodedata
	return unicodedata.normalize("NFKC", text).strip().upper()

def standardize_organism(name):
	
	import re
	name = name.strip()
	name = re.sub(r'\b([A-Z]{2,}[^\s]*|\d+)\b', '', name)
	name = re.sub(r'\s+', ' ', name).strip()
	return name

def search_genbank(email, query, retmax, usehistory=None, download=None):
	
	from Bio import Entrez

	Entrez.email = email
	with Entrez.esearch(db="nucleotide", term=query, retmax=retmax, usehistory=usehistory) as search_handle:

		search_results = Entrez.read(search_handle)

		if download:
			count = int(search_results["Count"])
			webenv = search_results["WebEnv"]
			query_key = search_results["QueryKey"]
			return count, webenv, query_key
		else: 
			return search_results["IdList"]

def count_genes(rec):
	genes_found = set()
	for feature in rec.features:
		if feature.type in ["gene", "CDS", "rRNA", "tRNA"]:
			for key in ["gene", "product", "note"]:
				if key in feature.qualifiers:
					genes_found.add(feature.qualifiers[key][0].strip().upper())
					break
	return len(genes_found)

def filter_ind_by_spp(all_records, alias_map, genes_selected, removal_reasons, unique, prioritize):
	
	from collections import defaultdict

	filtered_records = []
	groups_by_spp = defaultdict(list)

	for record in all_records:
		organism = record.annotations.get("organism", "")
		organism = standardize_organism(organism)
		voucher = ""
		isolate = ""
		for feature in record.features:
			if feature.type == "source":
				voucher = feature.qualifiers.get("specimen_voucher", [""])[0]
				isolate = feature.qualifiers.get("isolate", [""])[0]
				break
		ident = voucher or isolate or record.id
		groups_by_spp[(organism, ident)].append(record)

	candidates_by_spp = defaultdict(list)
	for (organism, ident), records in groups_by_spp.items():
		candidates_by_spp[organism].append((ident, records))

	for spp, candidates in candidates_by_spp.items():
		best = None
		best_id = None
		largest_n = -1
		for ident, records in candidates:
			genes_temp = set()
			seen_temp = set()
			dummy_dict = {}
			dummy_seen_leu_ser = set()
			dummy_leu = defaultdict(list)
			dummy_ser = defaultdict(list)
			for r in records:
				r.annotations["original_id"] = r.id.split(".")[0]
				extract_genes(
					r, "dummy", alias_map, genes_selected, seen_temp,
					dummy_dict, set(), set(), dummy_leu, dummy_ser, dummy_seen_leu_ser, unique, prioritize
				)
			genes_temp.update(dummy_dict.keys())
			if len(genes_temp) > largest_n:
				largest_n = len(genes_temp)
				best = records
				best_id = ident
			elif len(genes_temp) == largest_n and not best:
				best = records
				best_id = ident

		if best:
			filtered_records.extend(best)

		for ident, records in candidates:
			if ident != best_id:
				for r in records:
					removal_reasons[r.id].append("Removed by gene prioritization: another individual with more genes")

	return filtered_records

def filter_largest_fragment(gene_dict, header_to_species=None):

	from collections import defaultdict

	new_dict = {}
	for gene, entries in gene_dict.items():
		by_key = defaultdict(list)
		for header, seq, original_id in entries:
			key = (
				header_to_species.get(header)
				if header_to_species
				else header
			)
			if key:
				by_key[key].append((header, seq, original_id))

		new_dict[gene] = []
		for group in by_key.values():
			largest = max(group, key=lambda x: len(x[1]))
			new_dict[gene].append(largest)

	return new_dict

def fetch_with_retry(ids=None, rettype="gb", retmode="text", **kwargs):

	from http.client import IncompleteRead
	from Bio import Entrez, SeqIO
	import time

	global error_queue

	for try_n in range(3):
		try:
			if ids:
				with Entrez.efetch(db="nucleotide", id=ids, rettype=rettype, retmode=retmode, **kwargs) as handle:
					return list(SeqIO.parse(handle, "genbank"))
			else:
				with Entrez.efetch(db="nucleotide", rettype=rettype, retmode=retmode, **kwargs) as handle:
					return list(SeqIO.parse(handle, "genbank"))
		except IncompleteRead:
			report_message(f"IncompleteRead, try {try_n+1}/3...", "warning", "Warning")
			time.sleep(2)
		except Exception as e:
			report_message(f"Unexpected error: {e}", "error", "Error")
			error_queue.put(str(e))
			break
	return []
	
def get_standardized_organism(record):

	organism = record.annotations.get("organism", "")
	if not organism:
		for feature in record.features:
			if feature.type == "source" and "organism" in feature.qualifiers:
				organism = feature.qualifiers["organism"][0]
				break
	return standardize_organism(organism or "Unknown")

def classify_positional_tRNA(start, ref_pos_all, name_trna, tolerancia_bp=10):
	
	anchors = sorted(
		[(gene, pos) for gene, pos_list in ref_pos_all.items()
		 if not gene.upper().startswith("TRNA") for pos in pos_list],
		key=lambda x: x[1]
	)

	for gene, pos in anchors:
		if start < (pos + tolerancia_bp):
			return f"{name_trna}_before_start_{gene}"

	return f"{name_trna}_unclassified"

def extract_genes(record, header_id, alias_map, genes_selected, seen_genes, gene_dict, aliases_unknown, aliases_duplicated, grouper_leu, grouper_ser, seen_leu_ser, unique, prioritize, species_to_genes=None, organism=None):

	from collections import defaultdict

	original_id = record.annotations.get("original_id", record.id)

	alias_lookup = {}
	for main_name, aliases in alias_map.items():
		alias_lookup[main_name.upper()] = main_name
		for alias in aliases:
				alias_lookup[alias.upper()] = main_name

	genes_selected_set = {g.upper() for g in genes_selected} if genes_selected else None

	ref_pos = {}
	ref_pos_all = {}
	for feature in record.features:
		try:
			if feature.type not in ["gene", "CDS", "rRNA", "tRNA"]:
				continue

			raw_name = None
			for key in ["gene", "product", "note"]:
				if key in feature.qualifiers:
					raw_name = feature.qualifiers[key][0].strip()
					break
				if not raw_name:
					continue

			raw_name_upper = raw_name.upper()
			gene = None
			for main_name, aliases in alias_map.items():
				all_names = [main_name.upper()] + [a.upper() for a in aliases]
				if raw_name_upper in all_names:
					gene = main_name
					break
				
			if not gene:
				continue

			pos = int(feature.location.start)
			if gene not in ref_pos_all:
				ref_pos_all[gene] = []
			ref_pos_all[gene].append(pos)

		except:
			pass

	for feature in record.features:
		try:
			if feature.type not in ["gene", "CDS", "rRNA", "tRNA"]:
				continue

			raw_name = None
			for key in ["gene", "product", "note"]:
				if key in feature.qualifiers:
					raw_name = feature.qualifiers[key][0]
					break
			if not raw_name:
				continue

			raw_name = normalize(raw_name)
			raw_upper = raw_name.upper()
			name = alias_lookup.get(raw_upper)
			if not name:
				aliases_unknown[raw_name].add(record.id)
				continue

			if genes_selected_set and name.upper() not in genes_selected_set:
				continue

			if unique:
				if species_to_genes and organism:
					if name.upper() in species_to_genes.get(organism, set()):
						continue

			start = int(feature.location.start)
			end = int(feature.location.end)
			strand = feature.location.strand
			seq = record.seq[start:end]
			if strand == -1:
				seq = seq.reverse_complement()

			if name in ["tRNA-Leu", "tRNA-Ser"]:
				tag = classify_positional_tRNA(start, ref_pos_all, name)
				
				if unique:
					if species_to_genes and organism:
						if tag in species_to_genes.get(organism, set()):
							continue

						gene_name = tag.split("_")[0]
						tags_for_same_gene = [t for t in gene_dict if t.startswith(gene_name)]

						tags_with_header = set()
						for t in tags_for_same_gene:
							for h, _, oid in gene_dict[t]:
								if h == header_id:
									tags_with_header.add(t)

						# Agora permite até 2 tags diferentes com esse header_id
						if len(tags_with_header) >= 2 and tag not in tags_with_header:
							# header_id já está em pelo menos 2 tags diferentes e não está na tag atual,
							# então ignora essa entrada
							continue

				if tag not in gene_dict:
					gene_dict[tag] = []
				gene_dict[tag].append((header_id, str(seq), original_id))
				seen_genes.add((original_id, tag))

				key = (header_id, name, tag)
				if key not in seen_leu_ser:
					if name.upper() == "TRNA-LEU":
						grouper_leu[tag].append((header_id, str(seq), original_id))
					else:
						grouper_ser[tag].append((header_id, str(seq), original_id))
					seen_leu_ser.add(key)
				continue

			if (original_id, name) in seen_genes:
				aliases_duplicated.add(name)
				continue

			if name not in gene_dict:
				gene_dict[name] = []
			gene_dict[name].append((header_id, str(seq), original_id))
			seen_genes.add((original_id, name))
		except:
			pass

	#gene_dict.pop("tRNA-Leu", None)
	#gene_dict.pop("tRNA-Ser", None)

def overlap_animals(record, gene_dict, alias_map):
	
	original_id = record.annotations.get("original_id", record.id)
	name = record.annotations.get("organism", "Unknown")
	header_id = f"{original_id}"

	gene_coords = []
	for feature in record.features:
		if feature.type not in ["gene", "CDS", "rRNA"]:
			continue

		raw_name = None
		for key in ["gene", "product", "note"]:
			if key in feature.qualifiers:
				raw_name = feature.qualifiers[key][0]
				break

			if not raw_name:
				continue

			raw_upper = raw_name.upper()
			name = None
			for main_name, aliases in alias_map.items():
				if raw_upper == main_name.upper() or raw_upper in map(str.upper, aliases):
					name = main_name
					break

			if not name:
				continue

			start = int(feature.location.start)
			end = int(feature.location.end)
			strand = feature.location.strand

			gene_coords.append((name, start, end, strand, feature))

	gene_coords.sort(key=lambda x: x[1])
	log_records = []

	for i in range(len(gene_coords) - 1):
		name, start, end, strand, _ = gene_coords[i]
		next_name, next_start, next_end, _, _ = gene_coords[i + 1]

		if name != next_name and end > next_start:
			overlap = end - next_start
			log_records.append((name, start, end, next_name, next_start, next_end, overlap))

			new_entries = []
			for header, seq, orig_id in gene_dict.get(name, []):
				if orig_id != original_id:
					new_entries.append((header, seq, orig_id))
					continue

				if strand == 1:
					seq_corrected = seq[:next_start - start]
				else:
					cut = end - next_start
					seq_corrected = seq[cut:]

				new_entries.append((header, seq_corrected, orig_id))

			gene_dict[name] = new_entries

	return header_id, name, log_records
		
def metadata_by_gene(gene_dict, excel_data, fields_excel, output_dir, all_records, extract_all_genes):
	
	from openpyxl import Workbook
	import os
	
	wb = Workbook()

	ws = wb.active
	ws.title = "Metadata"

	header_map = dict(excel_data)

	ws.append(["GenBank_ID"] + fields_excel)
	excel_data.sort(key=lambda x: x[1].get("organism", "").lower())
	for rid, info in excel_data:
		ws.append([rid] + [info.get(f, "") for f in fields_excel])

	if extract_all_genes and len(gene_dict) > 1 and len(all_records) > 1:
		for gene, entries in sorted(gene_dict.items()):
			ws = wb.create_sheet(title=gene[:31])  # Excel limits tab name
			ws.append(["GenBank_ID"] + fields_excel)
				
			entries = sorted(entries, key=lambda x: header_map.get(x[2], {}).get("organism", "").lower())
				 
			for head, _, original_id in entries:
				meta = header_map.get(original_id, {}).copy()
				if "organism" in meta:
					meta["organism"] = standardize_organism(meta["organism"])
				ws.append([original_id] + [meta.get(f, "") for f in fields_excel])

	out_path = os.path.join(output_dir, "metadata.xlsx")
	wb.save(out_path)

def metadata_gene_vs_sample(gene_dict, excel_data, header_fields, output_dir, config):
	
	from openpyxl import Workbook
	from collections import defaultdict
	import os

	prioritize_more_genes = config.get("prioritize_more_genes", False)

	header_map = dict(excel_data)
	genes = sorted(gene_dict.keys())
	
	if config.get("prioritize_more_genes"):
		samples = defaultdict(lambda: {"genes": {}, "genbanks": set()})
		for gene, entries in gene_dict.items():
			for _, _, genbank_id in entries:
				meta = header_map.get(genbank_id, {})
				org = meta.get("organism", "").strip()
				org = standardize_organism(org)
				voucher = meta.get("specimen_voucher", "").strip()
				isolate = meta.get("isolate", "").strip()
				ident = voucher or isolate or genbank_id
				key = (org, ident)
				samples[key]["genes"][gene] = genbank_id
				samples[key]["metadata"] = {f: (org if f == "organism" else meta.get(f, '')) for f in header_fields}
				samples[key]["genbanks"].add(genbank_id)

		wb = Workbook()
		ws = wb.active
		ws.title = "metadata_gene_vs_sample"
		ws.append(header_fields + genes)

		ordered = sorted(
			samples.items(),
			key=lambda x: x[1]["metadata"].get("organism", "").lower()
		)

		for _, data in ordered:
			row = [data["metadata"].get(f, "") for f in header_fields]
			row += [data["genes"].get(g, "") for g in genes]
			ws.append(row)

	else:
		
		samples = {}
		for gene, seqs in gene_dict.items():
			for _, _, genbank_id in seqs:
				if genbank_id in header_map:
					if genbank_id not in samples:
						samples[genbank_id] = {
							"metadata": {f: header_map[genbank_id].get(f, '') for f in header_fields},
							"genes": {g: '' for g in genes}
						}
					samples[genbank_id]["genes"][gene] = genbank_id

		wb = Workbook()
		ws = wb.active
		ws.title = "metadata_gene_vs_sample"
		ws.append(header_fields + genes)

		ordered = sorted(
			samples,
			key=lambda x: standardize_organism(samples[x]["metadata"].get("organism", ""))
		)

		for genbank_id in ordered:
			meta = samples[genbank_id]["metadata"]
			gene_map = samples[genbank_id]["genes"]
			row = [meta[f] for f in header_fields] + [gene_map[g] for g in genes]
			ws.append(row)

	out_path = os.path.join(output_dir, "genes_matrix.xlsx")
	wb.save(out_path)

def make_query(email, alias_map, genes, organisms, mito, mitogenome, chloroplast, title, additional, min_len, max_len, add_unverified_exclusion):

	import logging

	# Generates gene synonyms
	gene_query = ""

	if isinstance(genes, list):
		selected_genes = [g.strip() for g in genes if g.strip()]
	elif isinstance(genes, str):
		selected_genes = [g.strip() for g in genes.split(",") if g.strip()]
	else:
		selected_genes = []
	grouped_aliases = []
	for g in selected_genes:
		match_found = False
		for main_name, aliases in alias_map.items():
			if g.upper() == main_name.upper():
				group = " OR ".join(f'"{a}"' if " " in a else a for a in aliases)
				grouped_aliases.append(f"({group})")
				match_found = True
				break
		if not match_found:
			grouped_aliases.append(f'"{g}"' if " " in g else g)
			
		if grouped_aliases:
			gene_query = " OR ".join(grouped_aliases)
		
	query_parts = []
	if gene_query:
		query_parts.append(f"({gene_query})")

	if organisms:
		org_query = " OR ".join([f'"{o.strip()}"[Organism]' for o in organisms.split(",")])
		query_parts.append(f"({org_query})")

	if mito:
		query_parts.append("mitochondrial")

	if mitogenome:
		query_parts.append("mitochondrion[Filter]")

	if chloroplast:
		query_parts.append("chloroplast[Filter]")

	if title:
		query_parts.append(title)
	 
	if min_len or max_len:
		min_val = int(min_len) if min_len else ""
		max_val = int(max_len) if max_len else ""
			
		if min_val and max_val:
			query_parts.append(f"{min_val}:{max_val}[SLEN]")
		elif min_val:
			query_parts.append(f"{min_val}:99999999[SLEN]")
		elif max_val:
			query_parts.append(f"2:{max_val}[SLEN]")

	query = " AND ".join(query_parts)
	query += " NOT wgs[prop] NOT tsa[prop]"
	if additional:
		query += f" {additional}"
	
	if add_unverified_exclusion:
		query += " NOT UNVERIFIED"

	try:
		ids = search_genbank(email, query, retmax=999999)
	except Exception as e:
		try:
			import tkinter
			from tkinter import messagebox
			root = tkinter._default_root
			if root is not None:
				report_message(f"An error occurred: {e}", "error", "Error", root)
			else:
				raise RuntimeError
		except:
			report_message(f"An error occurred: {e}", "error", "Error")
			ids = []
	print(query)
	return ids, query

def report_message(msg, level="info", title=None, root=None):

	import logging

	title = title or level.capitalize()
	if root is not None:
		try:
			import tkinter
			import tkinter.messagebox as messagebox
			if level == "error":
				messagebox.showerror(title, msg)
			elif level == "warning":
				messagebox.showwarning(title, msg)
			else:
				messagebox.showinfo(title, msg)
			return
		except Exception:
			pass 
	
	if level == "error":
		logging.error(f"{title}: {msg}")
	elif level == "warning":
		logging.warning(f"{title}: {msg}")
	else:
		logging.info(f"{title}: {msg}")

def sanitize_filename(filename):

	import re
	import platform
	
	if not filename:
		return "unnamed"
	
	invalid_chars = r'[<>:"|?*\\\/]'
	filename = re.sub(invalid_chars, '_', filename)
	
	if platform.system() == 'Windows':
		reserved = {'CON', 'PRN', 'AUX', 'NUL', 'COM1', 'COM2', 'COM3', 
				   'COM4', 'COM5', 'COM6', 'COM7', 'COM8', 'COM9',
				   'LPT1', 'LPT2', 'LPT3', 'LPT4', 'LPT5', 'LPT6', 
				   'LPT7', 'LPT8', 'LPT9'}
		
		name_only = filename.split('.')[0].upper()
		if name_only in reserved:
			filename = f"_{filename}"
	
	if len(filename) > 200:
		name, ext = filename.rsplit('.', 1) if '.' in filename else (filename, '')
		filename = name[:200-len(ext)-1] + ('.' + ext if ext else '')
	
	return filename.strip()

def ensure_writable_directory(path):

	import os
	
	try:
		os.makedirs(path, exist_ok=True)
		
		test_file = os.path.join(path, '.write_test')
		try:
			with open(test_file, 'w') as f:
				f.write('test')
			os.remove(test_file)
			return True, ""
		except Exception as e:
			return False, f"Cannot write to directory: {e}"
			
	except PermissionError:
		return False, "Permission denied to create directory"
	except Exception as e:
		return False, f"Cannot create directory: {e}"

def begin_search(config, root=None):

	from datetime import datetime
	import threading
	import logging
	import os
	import re

	logging.basicConfig(level=logging.INFO, format='%(levelname)s:%(message)s')

	from pynnotate.main.alias_maps import alias_map_animal, alias_map_mito_plant, alias_map_chloroplast

	report_message("WELCOME TO PYNNOTATE!\n", "info", "Info")

	email = config["email"]
	title = config["title"]
	additional = config["additional"]
	organisms = config["organisms"]
	genes = config["genes"]
	fields = config["fields"]
	output_name = config["output_name"]
	folder_name = config["folder_name"]
	if folder_name:
		folder_name = sanitize_filename(folder_name)
	include_id = config["include_id"]
	mito = config["mito"]
	mitogenome = config["mitogenome"]
	ids_text = config["ids_text"]
	chloroplast = config["chloroplast"]
	min_len = config["min_len"]
	max_len = config["max_len"]
	delete_unverified = config["delete_unverified"]
	extraction = config["extraction"]
	overlap = config["overlap"]
	unique = config["unique"]
	prioritize = config["prioritize_more_genes"]
	logmissing = config["logmissing"]
	org_type = config["org_type"]
	add_synonyms = config["add_synonyms"]

	if not extraction:
		if overlap:
			report_message("⚠️  '--overlap' ignored because '--extraction' was not enabled.", "warning", "Warning")
		if logmissing:
			report_message("⚠️  '--logmissing' ignored because '--extraction' was not enabled.", "warning", "Warning")
		overlap = False
	elif org_type != "animal_mito":
		if overlap:
			report_message("⚠️  '--overlap' is only applicable to 'animal_mito'. Ignoring.", "warning", "Warning")
		overlap = False

	alias_map_base = {
		"animal_mito": alias_map_animal,
		"plant_mito": alias_map_mito_plant,
		"plant_chloro": alias_map_chloroplast,
		"other": {}
	}.get(org_type, alias_map_animal)

	if add_synonyms:
		try:
			if isinstance(add_synonyms, dict):
				user_synonyms = add_synonyms
			elif isinstance(add_synonyms, str):
				user_synonyms = load_synonyms_from_str(add_synonyms, root)
			else:
				raise ValueError("add_synonyms must be a string or a dictionary.")
		except ValueError as e:
			report_message(f"Failed to parse synonyms: {e}", "error", "Error", root)
			return
	else:
		user_synonyms = {}

	alias_map = merge_synonyms(alias_map_base, user_synonyms)

	valid_fields = ["organism", "isolate", "specimen_voucher", "country",
		"collection_date", "host", "strain", "isolation_source", "lat_lon"]

	add_unverified_deletion = delete_unverified

	if not fields or not output_name or not email:
		report_message("Fill in the required fields: email, output, and header", "warning", "Required fields", root)
		return

	if isinstance(fields, list):
		header_fields = [i.strip() for i in fields if i.strip()]
	elif isinstance(fields, str):
		header_fields = [i.strip() for i in re.split(r"[\s,]+", fields) if i.strip()]
	else:
		header_fields = []

	invalid_fields = [f for f in header_fields if f not in valid_fields]
	
	if invalid_fields:
		report_message(f"Invalid fields: {', '.join(invalid_fields)}", "error", "Error", root)
		return

	if min_len or max_len:
		try:
			min_val = int(min_len) if min_len else ""
			max_val = int(max_len) if max_len else ""
		except ValueError:
			report_message("Min/max size values must be integers.", "error", "Error", root)
			return

	date_string = datetime.now().strftime("%d%b%y").lower()
	time_string = datetime.now().strftime("%Hh%Mm%Ss").lower()
	folder = output_name if isinstance(output_name, str) else output_name[0]

	can_write, error_msg = ensure_writable_directory(folder)
	if not can_write:
		report_message(f"Cannot create output directory: {error_msg}", "error", "Error", root)
		return

	if folder_name:
		output_dir = os.path.join(folder, folder_name)
	else:
		output_dir = os.path.join(folder, f"pynnotate_{date_string}_{time_string}")

	os.makedirs(output_dir, exist_ok=True)

	if ids_text:
		if root is not None:
			from tkinter import messagebox
			ids = [i.strip() for i in re.split(r"[\s,]+", ids_text) if i.strip()]
			proceed_man = messagebox.askyesno("Confirmation", f"🔢 {len(ids)} IDs provided manually.\n\nDo you want to start the download?")
		else:
			ids = ids_text
			response_man = input(f"🔢 {len(ids)} IDs provided manually.\n\nDo you want to start the download? (y/n): ")
			proceed_man = response_man.lower() in ["y", "yes"]

		if proceed_man:
			config = {
				"include_id": include_id,
				"extract_all_genes": extraction,
				"delete_overlap": overlap,
				"unique_species": unique,
				"prioritize_more_genes": prioritize,
				"min_len": int(min_len or 0),
				"max_len": int(max_len or 999999),
				"log_missing": logmissing,
				"org_type": org_type,
				"add_synonyms": add_synonyms,
				"mitogenome": mitogenome,
				"chloroplast": chloroplast
			}
			threading.Thread(
				target=download_sequences,
				args=(email, ids, output_dir, header_fields, alias_map, config, "", None, root)
			).start()
		return

	if not any([genes, organisms, title, mito, mitogenome]):
		report_message("Invalid search: Please enter genes, organism, title or select at least one mitochondria/mitogenome option to perform the search.", "warning", "Warning", root)
		if root:
			root.after(0, lambda: btn.config(text="💾 Search and download sequences", state="normal"))
			root.after(0, lambda: set_widgets_state(True))
		return

	ids, query = make_query(email, alias_map, genes, organisms, mito, mitogenome, chloroplast, title, additional, min_len, max_len, add_unverified_deletion)

	if ids:
		if root is not None:
			from tkinter import messagebox
			proceed_down = messagebox.askyesno("Records found", f"🔢 {len(ids)} records found.\n\nDo you want to start downloading?")
		else:
			response_down = input(f"🔢 {len(ids)} records found.\n\nDo you want to start downloading? (y/n): ")
			proceed_down = response_down.lower() in ["y", "yes"]
		if proceed_down:
			config = {
				"include_id": include_id,
				"extract_all_genes": extraction,
				"delete_overlap": overlap,
				"unique_species": unique,
				"prioritize_more_genes": prioritize,
				"min_len": int(min_len or 0),
				"max_len": int(max_len or 999999),
				"log_missing": logmissing,
				"org_type": org_type,
				"add_synonyms": add_synonyms,
				"mitogenome": mitogenome,
				"chloroplast": chloroplast
			}
			threading.Thread(
				target=download_sequences,
				args=(email, ids, output_dir, header_fields, alias_map, config, genes, query, root)
			).start()
	else:
		report_message("No results: No ID found for search.", "info", "Info", root)

def download_sequences(email, genbank_ids, output_dir, header_fields, alias_map, config, genes, query=None, root=None):

	import os
	from tqdm import tqdm
	from Bio import Entrez, SeqIO
	from openpyxl import Workbook
	from collections import defaultdict
	import traceback

	from pynnotate.graphic.utils import create_loading_window_with_progress

	Entrez.email = email

	valid_fields = ["organism", "isolate", "specimen_voucher", "country",
		"collection_date", "host", "strain", "isolation_source", "lat_lon"]

	genbank_ids = genbank_ids or []
	include_id = config["include_id"]
	extract_all_genes = config["extract_all_genes"]
	delete_overlap = config["delete_overlap"]
	unique_species = config.get("unique_species", False)
	prioritize_more_genes = config.get("prioritize_more_genes", False)
	log_missing = config.get("log_missing", False)
	org_type = config["org_type"]
	add_synonyms = config["add_synonyms"]
	mitogenome = config["mitogenome"]
	chloroplast = config["chloroplast"]

	if not genbank_ids and not query:
		report_message("No ID found.", "warning", "Warning", root)
		return
	
	if isinstance(genes, str):
		genes_list = [g.strip() for g in genes.split(",") if g.strip()]
	elif isinstance(genes, list):
		genes_list = [g.strip() for g in genes if isinstance(g, str) and g.strip()]
	else:
		genes_list = []

	genes_selected = set()

	for g in genes_list:
		norm_g = normalize(g)
		match = None
		for main_name in alias_map:
			if normalize(main_name) == norm_g:
				match = main_name
				break
		genes_selected.add(match if match else g)
		
	missing_by_record = defaultdict(list) if log_missing else {}
	
	removal_reasons = defaultdict(list)

	if extract_all_genes and not genes_selected:
		if mitogenome:
			all_keys = set(alias_map.keys())
			keys_to_remove = {"18S", "28S", "ITS1", "ITS2", "RAG1", "RAG2", "BDNF", "CMOS", "POMC", "TYR", "RHO", "SIAH1", "CRYB", "TNS3", "S7", "TYRP1", "EIF3", "CXCR4", "NT3", "FIB7", "MYH6", "CMYC", "CMYC2", "H3A", "ZEB2", "MC1R", "SNCAIP", "PTPRC", "OPA1"}
			genes_selected = all_keys - keys_to_remove
			genes_selected_original = genes_selected
			report_message("No genes specified, but mitogenome flag enabled — extracting all mitogenome genes", "info", "Info", root)
		else:
			genes_selected_original = genes_selected
			genes_selected = set(alias_map.keys())
			report_message("No genes specified — extracting all genes", "info", "Info", root)

	unknown_aliases = defaultdict(set)
	duplicated_aliases = set()
	grouper_leu = defaultdict(list)
	grouper_ser = defaultdict(list)

	try:
		os.makedirs(output_dir, exist_ok=True)

		can_write, error_msg = ensure_writable_directory(output_dir)
		if not can_write:
			report_message(f"Output directory problem: {error_msg}", "error", "Error", root)
			return

		gene_folder = os.path.join(output_dir, "genes")
		fasta_file = os.path.join(output_dir, "sequences.fasta")
		excel_file = os.path.join(output_dir, "metadata.xlsx")
		log_file = os.path.join(output_dir, "log.txt")

		if root is not None:
			loading_win = None
			loading_win, progress_var, progress_bar = create_loading_window_with_progress(root)

		all_records = []
		batch_size = 500
		ignored_ids = []

		pbar=None
		if query:
			
			count, webenv, query_key = search_genbank(email=email, query=query, retmax=0, usehistory="y", download=True)

			if root is None:
				pbar = tqdm(total=count, desc="Downloading data from GenBank... Please wait a few minutes", unit="seq")

			for start in range(0, count, batch_size):

				records = fetch_with_retry(
					retstart=start, retmax=batch_size,
					webenv=webenv, query_key=query_key
				)

				for rec in records:
					try:
						seq_string = str(rec.seq).upper()
						if not rec.seq or len(seq_string.strip("N")) == 0 or "SEQUENCE CONTENT IS UNDEFINED" in seq_string:
							ignored_ids.append(rec.id)
							continue
						all_records.append(rec)
					except Exception:
						ignored_ids.append(rec.id)
						continue
				if root is not None:
					progress = int(((start + batch_size) / count) * 100)
					progress_var.set(min(progress, 100))
					progress_bar.update()
					root.update()
				else:
					pbar.update(batch_size)

		else:

			if root is None:
				pbar = tqdm(total=len(genbank_ids), desc="Downloading data from GenBank... Please wait a few minutes", unit="seq")
			
			for i in range(0, len(genbank_ids), batch_size):
				chunk = genbank_ids[i:i + batch_size]
				records = fetch_with_retry(ids=chunk)
				for rec in records:
					try:
						seq_string = str(rec.seq).upper()
						if not rec.seq or len(seq_string.strip("N")) == 0 or "SEQUENCE CONTENT IS UNDEFINED" in seq_string:
							ignored_ids.append(rec.id)
							continue
						all_records.append(rec)
					except Exception:
						ignored_ids.append(rec.id)
						continue
				if root is not None:
					progress = int(((i + batch_size) / len(genbank_ids)) * 100)
					progress_var.set(min(progress, 100))
					progress_bar.update()
					root.update()
				else:
					pbar.update(batch_size)

		all_records_orig = all_records.copy()
		min_len = config.get("min_len", 0)
		max_len = config.get("max_len", 999999)

		all_records = [rec for rec in all_records if min_len <= len(rec.seq) <= max_len]

		if not all_records:
			report_message(f"No sequences kept after aplying length filter: {min_len}-{max_len}.", "error", "Error", root)
			return

		use_genes_priority = False

		if prioritize_more_genes:
			if not genes.strip():
				use_genes_priority = True
			else:
				num_genes = len([g for g in genes.split(",") if g.strip()])
				if num_genes > 1:
					use_genes_priority = True
	
		if use_genes_priority:
			all_records = filter_ind_by_spp(
				all_records, alias_map, genes_selected, removal_reasons, unique_species, prioritize_more_genes
			)

		if unique_species:

			records_by_species = defaultdict(list)
			for rec in all_records:
				organism = rec.annotations.get("organism", "")
				if not organism:
					for feature in rec.features:
						if feature.type == "source" and "organism" in feature.qualifiers:
							organism = feature.qualifiers["organism"][0]
							break
				if not organism:
					organism = "Unknown"
				organism = standardize_organism(organism)
				records_by_species[organism].append(rec)

			all_records = []
			for spp, recs in records_by_species.items():
				recs_sorted = sorted(recs, key=count_genes, reverse=True)
				all_records.extend(recs_sorted)

		filtered_records = []
		species_seen = set()
		seen_headers = set()
		excel_data = []
		fields_excel = valid_fields
		gene_dict = {}
		original_id_to_species = {}
		ignored_aliases = set()
		all_log_overlap = []
		seen_genes = set()
		seen_headers = set()
		seen_leu_ser = set()

		species_to_genes = defaultdict(set)

		for record in all_records:

			genbank_id = record.id.split(".")[0]
			record.annotations["original_id"] = genbank_id
			
			organism = record.annotations.get("organism", "")
			if not organism:
				for feature in record.features:
					if feature.type == "source" and "organism" in feature.qualifiers:
						organism = feature.qualifiers["organism"][0]
						break
			if not organism:
				organism = "Unknown"

			organism = standardize_organism(organism)

			current_record_genes = set()
			for feature in record.features:
				if feature.type in ["gene", "CDS", "rRNA", "tRNA"]:
					gene_name = None
					for key in ["gene", "product", "note"]:
						if key in feature.qualifiers:
							gene_name = feature.qualifiers[key][0].strip()
							break
						if not gene_name:
							continue
					if gene_name:
						gene_name_upper = gene_name.upper()
						current_record_genes.add(gene_name_upper)

			if unique_species:
				new_genes = current_record_genes - species_to_genes[organism]
				if not new_genes:
					removal_reasons[genbank_id].append("All genes already saved for species (flexible mode⎯unique_species enabled)")
					continue

			if prioritize_more_genes:
				if organism in species_seen:
					removal_reasons[genbank_id].append("Another sequence for species already added (strict mode⎯prioritize enabled)")
					continue
			species_seen.add(organism)

			values = []
			feature_info = {f: "" for f in fields_excel}
			for field in header_fields:
				val = ""
				for feature in record.features:
					if feature.type == "source":
						val = feature.qualifiers.get(field, [""])[0].strip()
						break
				values.append(val.replace(" ", "_") if val else "")
		
			for feature in record.features:
				if feature.type == "source":
					for f in fields_excel:
						if f in feature.qualifiers:
							feature_info[f] = feature.qualifiers[f][0]
					break
		
			base_header = "_".join([v for v in values if v]).strip("_")
			if not base_header:
				base_header = organism.replace(" ", "_")
			header_id = f"{base_header}_{genbank_id}" if include_id else base_header
			original_header = header_id
				
			if not prioritize_more_genes:
				suffix = 1
				while header_id in seen_headers:
					header_id = f"{original_header}_{genbank_id if include_id else suffix}"
					suffix += 1
				seen_headers.add(header_id)
			else:
				seen_headers.add(header_id)
		
			if not include_id and genbank_id in header_id:
				header_id = header_id.replace(f"_{genbank_id}", "")
		
			record.id = header_id
			record.description = ""

			original_id_to_species[header_id] = organism

			genes_found = set()

			overlap_header=None
			overlap_name=None
			overlap_records=[]

			if extract_all_genes:

				extract_genes(
					record,
					header_id,
					alias_map,
					genes_selected,
					seen_genes,
					gene_dict,
					unknown_aliases,
					duplicated_aliases,
					grouper_leu,
					grouper_ser,
					seen_leu_ser,
					unique_species,
					prioritize_more_genes,
					species_to_genes,
					organism
				)
				
				if delete_overlap and org_type == "animal_mito":
					overlap_header, overlap_name, overlap_records = overlap_animals(record, gene_dict, alias_map)

				for gene, seqs in gene_dict.items():
					for head, seq, seq_original_id in seqs:
						if seq_original_id == record.annotations["original_id"]:
							genes_found.add(gene.upper())
							break

			species_to_genes[organism].update(genes_found)

			if not extract_all_genes or genes_found:
				filtered_records.append(record)
				excel_data.append((record.annotations["original_id"], feature_info))
			else:
				removal_reasons[genbank_id].append("Gene extraction failed: Unwanted gene according to the parameters specified")
				continue

		if prioritize_more_genes or unique_species:
			gene_dict = filter_largest_fragment(gene_dict)
		
		if unique_species:
			gene_dict = filter_largest_fragment(gene_dict, original_id_to_species)
		
		if extract_all_genes and not genes:
			genes_selected = set(gene_dict.keys())

		if filtered_records:
			with open(fasta_file, "w", encoding='utf-8', newline='\n') as f_out:
				SeqIO.write(filtered_records, f_out, "fasta")

		for tag, entries in grouper_leu.items():
			os.makedirs(gene_folder, exist_ok=True)
			path = os.path.join(gene_folder, f"{tag}.fasta")
			with open(path, "w", encoding='utf-8', newline='\n') as out:
				for header, seq, oid in entries:
					out.write(f">{header}\n{seq}\n")
		 
		for tag, entries in grouper_ser.items():
			os.makedirs(gene_folder, exist_ok=True)
			path = os.path.join(gene_folder, f"{tag}.fasta")
			with open(path, "w", encoding='utf-8', newline='\n') as out:
				for header, seq, oid in entries:
					out.write(f">{header}\n{seq}\n")

		with open(log_file, "w", encoding='utf-8', newline='\n') as lf:
			lf.write("PYNNOTATE\n")
			lf.write("\n")
			lf.write(f"User email: {email}\n")
			lf.write(f"Output folder: {output_dir}\n")
			lf.write("\n")
			lf.write(f"FASTA file generated: {fasta_file}\n")
			if extract_all_genes:
				lf.write(f"FASTA files separated by genes generated: {gene_folder}\n")
			lf.write(f"Generated Excel file: {excel_file}\n")
			lf.write("\n")
			if query:
				lf.write(f"Search query used: {query}\n")
			else:
				lf.write(f"GenBank IDS search: {genbank_ids}\n")
			lf.write(f"Genome type chosen: {org_type}\n")
			if add_synonyms:
				lf.write(f"New synonyms given!\n")
			lf.write("\n")

			if not unique_species and not prioritize_more_genes:
				lf.write("[🌐 UNCONSTRAINED MODE: INCLUDING ALL SEQUENCES]\n")
			elif unique_species:
				lf.write("[🌱 FLEXIBLE MODE ENABLED: ALLOWING MULTIPLE SEQUENCES PER SPECIES IF THEY HAVE NEW GENES]\n")
			elif prioritize_more_genes:
				lf.write("[🔒 STRICT MODE: INCLUDING ONLY ONE SEQUENCE PER SPECIES, PRIORITIZING INDIVIDUALS WITH MORE GENES]\n")
			lf.write("\n")

			lf.write(f"Total sequences found (before filters): {len(all_records_orig)}\n")
			lf.write(f"Total sequences saved: {len(filtered_records)}\n")
			lf.write(f"Total species saved: {len(species_seen)}\n")
			lf.write(f"Fields used in the header: {', '.join(header_fields)}\n")
			lf.write(f"Include GenBank ID in header: {'Yes' if include_id else 'No'}\n")
			lf.write("\n")

			if extract_all_genes:
				lf.write(f"[🌟 SEPARATE GENE EXTRACTION ENABLED]: {len(gene_dict)} genes grouped by synonym\n")

			if extract_all_genes and gene_dict:
				lf.write("[EXTRACT GENES SUMMARY]:\n")
				for gene in sorted(gene_dict):
					entries = gene_dict[gene]
					os.makedirs(gene_folder, exist_ok=True)
					safe_gene_name = sanitize_filename(gene)
					path = os.path.join(gene_folder, f"{safe_gene_name}.fasta")
					with open(path, "w", encoding='utf-8', newline='\n') as out:
						for head, seq, *_ in entries:
							out.write(f">{head}\n{seq}\n")
					lf.write(f"{gene}: {len(entries)} extracted sequences\n")

				grouper_combined = list(grouper_leu.items()) + list(grouper_ser.items())
				lf.write("\n[TRNA-LEU/SER CLUSTERINGS BY POSITIONAL PATTERN]: When using flexible mode, it's expected that some species will have more than two copies of tRNA-Leu and tRNA-Ser. This happens because GenBank entries often lack a clear specification about which copy of Leu/Ser a sequence refers to. Due to this, pynnotate strategy is that genes are extracted based on their genomic position. Therefore, it's up to the researcher to manually inspect and process these cases afterwards.\n")
				
				tags_set = sorted({tag for tag, _ in grouper_combined})
				
				for tag in tags_set:
					count = 0
					for name in gene_dict:
						if name == tag:
							for header, seq, oid in gene_dict[name]:
								count += 1
					lf.write(f"{tag}: {count} grouped sequences\n")
				lf.write("\n")

			if delete_overlap and org_type == "animal_mito":
				if not overlap_records:
					lf.write("[FIXING OVERLAP BETWEEN EXTRACTED GENES ENABLED] No overlap found!\n")
				else:
					lf.write("[FIXING OVERLAP BETWEEN EXTRACTED GENES ENABLED] Here are the bp removed:\n")
					key = f"{overlap_name} ({overlap_header})"
					lf.write(f"[{key}]\n")
					for g1, s1, e1, g2, s2, e2, o in overlap_records:
						lf.write(f"⚠️ {g1} ({s1}..{e1}) overlap {g2} ({s2}..{e2}): {o} bp removed from the end of {g1}\n")
				lf.write("\n")

			if unknown_aliases:
				lf.write("[UNKNOWN ALIASES] Some aliases were not detected in the synonyms dictionary and were not extracted:\n")
				for alias in sorted(unknown_aliases):
					origins = ", ".join(sorted(unknown_aliases[alias]))
					lf.write(f"{alias} (from records: {origins})\n")
				lf.write("\n")

			if duplicated_aliases:
				lf.write("[DUPLICATED ALIASES] Duplicated aliases are gene names that appear more than once for the same individual or species. These duplicates are ignored to avoid ambiguity during extraction:\n")
				for alias in sorted(duplicated_aliases):
					lf.write(alias + "\n")
				lf.write("\n")

			if log_missing and genes_selected_original:
				lf.write("[MISSING GENES] Some records did not present the genes of interest:\n")
				for record in filtered_records:
					found_genes = set()
					record_id = record.annotations.get("original_id", record.id)
					for gene, seqs in gene_dict.items():
						for _, _, seq_original_id in seqs:
							if seq_original_id == record_id:
								found_genes.add(gene)
								break
					missing = genes_selected_original - found_genes
					if missing:
						name = record.annotations.get("organism", "Unknown")
						chave = f"{name} ({record_id})"
						lf.write(f"[{chave}]\n")
						for gene in sorted(missing):
							lf.write(f"⚠️ {gene} missing\n")
						lf.write("\n")

			if removal_reasons:
				lf.write(f"[🔎 SEQUENCES REMOVED BY APPLIED FILTERS]: {len(all_records_orig) - len(filtered_records)}\n")
				all_reasons = defaultdict(list)
				for rec_id, reasons in removal_reasons.items():
					for reason in reasons:
						all_reasons[reason].append(rec_id)
						
				for reason, ids in all_reasons.items():
					ids_ordered = sorted(set(ids))
					lf.write(f"{reason}: {', '.join(ids_ordered)}\n")

			else:
				lf.write("✅ No sequences were removed by filters.\n")
			lf.write("\n")

		if root is not None:
			loading_win.destroy()
		else:
			pbar.close()

		report_message(f"Download complete!", "info", "Success", root)
	
		metadata_by_gene(gene_dict, excel_data, fields_excel, output_dir, all_records, extract_all_genes)

		if extract_all_genes and len(gene_dict) > 1 and len(all_records) > 1:
			metadata_gene_vs_sample(gene_dict, excel_data, header_fields, output_dir, config)

	except Exception as e:
		if root is not None:
			if loading_win:
				loading_win.destroy()
		else:
			if pbar:
				pbar.close()
		traceback.print_exc()
		report_message(f"Error processing sequences: {e}", "error", "Error", root)
